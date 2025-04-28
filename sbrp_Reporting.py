from venv import create
from openpyxl import load_workbook, Workbook  # 合并
import glob  # 合并
import pandas as pd  # 比对 & 处理数据
import datetime
import os
import nltk
import shutil  # file operation
import shutilwhich  # file operation

# ---变量
path_tmp = "/Users/linyunxia/Documents/report/tmp"
pathRMH = "/Users/linyunxia/Documents/report/org"
path_folder = "/Users/linyunxia/Documents/report/result/"
day_folder = datetime.datetime.now().strftime('%Y') + "/" + datetime.datetime.now().strftime(
    '%m') + "/" + datetime.datetime.now().strftime('%d')
time_folder = datetime.datetime.now().strftime("%H")
new_workbook = Workbook()
new_sheet = new_workbook.active
mkdir_target_path = "/Users/linyunxia/Documents/report/result" + day_folder
new_target_path = mkdir_target_path + "/" + time_folder

# -- 新建文件夹
if not os.path.exists(new_target_path):
    os.makedirs(new_target_path)

# 用flag变量明确新表是否已经添加了表头，只要添加过一次就无须重复再添加
# 0 or 1
flag = 1

for file in glob.glob(path_tmp + '/*.xlsx'):
    workbook = load_workbook(file)
    sheet = workbook.active

    coloum_A = sheet['A']
    row_lst = []
    for cell in coloum_A:
        if cell:
            print(cell.row)
            row_lst.append(cell.row)

    if not flag:
        header = sheet[1]
        header_lst = []
        for cell in header:
            header_lst.append(cell.value)
        new_sheet.append(header_lst)
        flag = 1

    for row in row_lst:
        data_lst = []
        for cell in sheet[row]:
            data_lst.append(cell.value)
        new_sheet.append(data_lst)

new_workbook.save(path_tmp + '/' + 'temp.xlsx')

# ---筛选比对 vlookup
# 总表
df_temp = pd.read_excel(path_tmp + '/' + 'temp.xlsx', sheet_name='Sheet')
df_RPA = pd.read_excel(pathRMH + '/' + 'GC Hotel System List 2024.xlsx', sheet_name='RPA SBRP')
df_Details = pd.read_excel(pathRMH + '/' + 'GC Hotel System List 2024.xlsx', sheet_name='Details')
data_temp = df_temp.merge(df_Details, on=['Inncode'])
data_final = data_temp.merge(df_RPA, on=['Inncode'])
data_final.to_csv(path_tmp + '/' + 'final.csv', index=False, encoding='utf_8_sig')

# 所有Oasis系统酒店分表
data_final_oasis_temp = data_final.loc[data_final['PMS_x'] == "Oasis"]
data_final_oasis = data_final_oasis_temp.iloc[:, [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 28, 29, 89]]
data_final_oasis.to_csv(path_tmp + '/' + 'final_oasis.csv', index=False, encoding='utf_8_sig')

# 所有Opera系统酒店分表
data_final_opera_temp = data_final.loc[data_final['PMS_x'] == "Opera"]
data_final_opera = data_final_opera_temp.iloc[:, [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 27, 28, 88]]
data_final_opera.to_csv(path_tmp + '/' + 'final_opera.csv', index=False, encoding='utf_8_sig')

# 所有SEP系统酒店分表
data_final_sep_temp = data_final.loc[data_final['PMS_x'] == "SEP"]
data_final_sep = data_final_sep_temp.iloc[:, [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 27, 28, 88]]
data_final_sep.to_csv(path_tmp + '/' + 'final_sep.csv', index=False, encoding='utf_8_sig')

# 所有HIEX酒店分表
data_final_HIEX_temp = data_final.loc[data_final['Sub Region'] == "HIEX"]
data_final_HIEX = data_final_HIEX_temp.iloc[:, [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 27, 28, 88]]
data_final_HIEX.to_csv(path_tmp + '/' + 'final_HIEX.csv', index=False, encoding='utf_8_sig')

# 所有Full Service酒店分表
data_final_Full_temp = data_final.loc[data_final['Sub Region'] != "HIEX"]
data_final_Full = data_final_Full_temp.iloc[:, [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 27, 28, 88]]
data_final_Full.to_csv(path_tmp + '/' + 'final_Full.csv', index=False, encoding='utf_8_sig')

# 截取行数(Error)
data_final_full_error_temp1 = data_final.loc[(data_final['OTB_Uploaded'] != "Completed")]
data_final_full_error_temp2 = data_final_full_error_temp1.iloc[:, [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]]
data_final_full_error = data_final_full_error_temp2.replace(
    ["The OTB file you are attempting to upload is out of chronological sequence."], ["OTB历史文件有遗漏请手工上传"])
data_final_full_error.to_csv(path_tmp + '/' + 'final_Full_error.csv', index=False, encoding='utf_8_sig')

# 所有Opening酒店分表
data_final_HIEX_temp = data_final.loc[data_final['Opening Status_x'] == "Open - Accepting Guests"]
data_final_HIEX = data_final_HIEX_temp.iloc[:, [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 28, 29, 77]]
data_final_HIEX.to_csv(path_tmp + '/' + 'final_opening.csv', index=False, encoding='utf_8_sig')


# ---移动文件
def move_file(old_path, new_path):
    print(old_path)
    print(new_path)
    filelist = os.listdir(old_path)
    print(filelist)
    for file in filelist:
        src = os.path.join(old_path, file)
        dst = os.path.join(new_path, file)
        print('src:', src)
        print('dst:', dst)
        shutil.move(src, dst)


if __name__ == '__main__':
    move_file(path_tmp, new_target_path)